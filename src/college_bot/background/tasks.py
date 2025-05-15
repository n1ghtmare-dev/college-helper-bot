from utils.json_handler.data_manager import JsonDataManager
from config import settings
from services.crud.groups_crud import get_schedule
from datetime import datetime, timedelta
from core.dispatcher import bot
from keyboards.inline import get_report_kb
from pathlib import Path
from openpyxl.drawing.image import Image
from openpyxl.utils import get_column_letter
import pandas as pd
from aiogram.types.input_file import FSInputFile
import json
import logging
import os



logger = logging.getLogger(__name__)

class Tasks:
    def __init__(self):
        self.json_manager = JsonDataManager(settings.DB_CONFIG)

    def update_json_from_db(self):
        date = datetime.now().strftime("%d-%m-%Y")
        schedule = get_schedule(date)

        groups_data = self.json_manager.get_groups_data()

        for item in schedule:
            if item['group_id'] not in groups_data:
                groups_data[item['group_id']] = {
                    "Количество": 0,
                    "Старосты": [],
                    "Расписание": []
                }

            groups_data[item['group_id']]['Расписание'] = [{
                'time': item['time'],
                'teacher_id': item['teacher_id']
            }]
            
        self.json_manager.save_to_json(groups_data)
        
    def json_update(self):
        logger.info("Update groups.json from database - started")
        self.update_json_from_db()
        logger.info("groups.json was successfully updated")

    async def create_report(self):

        all_headmen = self.json_manager.get_all_headmen()
        whole_schedule_data = self.json_manager.get_groups_data()

        for group_id, group_info in whole_schedule_data.items():
            if 'Старосты' in group_info and len(group_info['Старосты']) < 1:
                continue
            else:
                if 'Расписание' in group_info and len(group_info['Расписание']) < 1:
                    continue
                else:
                    for lesson in group_info['Расписание']:
                        print("TIME - ", lesson['time'])
                        print("ПРОВЕРКА РАЗНИЦЫ ВО ВРЕМЕНИ")
                        if self.check_time(lesson['time']):
                            # TODO: -> SEND MESSAGE
                            print("ОТПРАВКА СООБЩЕНИЯ - ", group_info['Старосты'][0])
        # FOR TEST REPORT MESSAGE
        await bot.send_message(5438186408, 'Нажмите для составления отчёта:', reply_markup=get_report_kb())

    def check_time(self, lesson_start_time: str) -> bool:
        current_time = datetime.now().time()

        start_str, end_str = lesson_start_time.split('-')
        start_time = datetime.strptime(start_str, '%H:%M').time()

        start_datetime = datetime.combine(datetime.today(), start_time)

        start_with_buffer = (datetime.combine(datetime.today(), start_time) + 
                             timedelta(minutes=20)).time()
        
        end_buffer = (start_datetime + timedelta(minutes=21, seconds=30)).time()
        
        print(f"TIME ZONE: {start_with_buffer} <= {current_time} < {end_buffer}")
        
        if start_with_buffer <= current_time and current_time < end_buffer:
            return True
        return False
    
    async def send_excel(self) -> bool:
        user_id = 5438186408
        try:
            excel_file = self.create_excel()
        except Exception as e:
            logger.error(f"Не удалось сформировать excel файл, проверьте существующие отчеты!, E - {e}")

        if not os.path.exists(excel_file):
            await bot.send_message(user_id, "Ошибка: не удалось создать файл отчета")
            return
        
        print("Отчет сформирован")

        await bot.send_document(
            chat_id=user_id,
            document=FSInputFile(excel_file, filename="report_summary.xlsx"),
            caption="Ваш отчет готов! 📊"
        )
        print("Отчет отправлен")

    def create_excel(self) -> str:

        reports_dir = Path(__file__).resolve().parent.parent.parent.parent / "data/reports"
        print(reports_dir)

        reports = []

        for report_file in  reports_dir.glob("*.json"):
            with open(report_file, 'r', encoding='utf-8') as f:
                reports.append(json.load(f))

        df = pd.DataFrame(reports)
        print(df)
    
        df['Дата'] = pd.to_datetime(df['Дата'])
        df['Дата'] = df['Дата'].dt.date

        excel_path = reports_dir.parent / "reports_summary.xlsx"
        with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Отчеты')

            workbook = writer.book
            worksheet = workbook['Отчеты']
            
            photo_col_idx = len(df.columns) + 1
            worksheet.cell(row=1, column=photo_col_idx, value="Фото")
            worksheet.column_dimensions[get_column_letter(photo_col_idx)].width = 20
            
            for idx, row in enumerate(df.itertuples(), start=2):
                if hasattr(row, 'photo_path') and row.photo_path:
                    try:
                        print(row.photo_path)
                        img = Image(row.photo_path)
                        img.width = 100
                        img.height = 100
                        worksheet.add_image(img, f"{get_column_letter(photo_col_idx)}{idx}")
                        worksheet.row_dimensions[idx].height = 80
                    except Exception as e:
                        print(f"Не удалось вставить изображение {row.photo_path}: {e}")
                        worksheet.cell(row=idx, column=photo_col_idx, value="Ошибка загрузки изображения")
            
            summary = df.pivot_table(
                index='Дата',
                values=['Количество студентов', 'Присутствие преподавателя'],
                aggfunc={'Количество студентов': 'sum', 'Присутствие преподавателя': 'mean'}
            )
            summary.to_excel(writer, sheet_name='Сводка')

        # Удаляем исходные файлы
        # for report_file in reports_dir.glob("*.json"):
        #     report_file.unlink()

        # for image_file in IMAGES_DIR.glob("*.jpg"):
        #     image_file.unlink()

        return excel_path



